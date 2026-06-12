#!/usr/bin/env python3
"""
build_labeled_variant_set.py
============================
把 VWD 分型突变库(/Volumes/LQ1000/VWD: 每型一个 xlsx + 部分 json)整理成**干净的
标签变体清单**, 同时喂两条线:
  - MD 扩容: A1 域内、单标签的 2B/2M 变体 → build_2b_mutants_foldx.py --variants-file
  - 分类器校准: 更大的 known 2B/2M 标签集

做的事:
  1. 解析每个分型文件夹(xlsx 三字母 + json 单字母)→ 变体, label = 文件夹分型
  2. 三字母→单字母; 按位置标注 domain + 是否落 7A6O 可 graft 的 A1 域(1262-1466)
  3. **跨表去重 + 冲突检测**: 同一变体出现在 ≥2 个分型表 → conflict(多效性/报道歧义)
  4. 输出 CSV + 冲突报告 + FoldX 直接可用的变体清单

输出(默认 repo output/):
  labeled_variants_all.csv        — 全部变体: aa_change, label(s), domain, in_a1_7a6o, n_labels, conflict, sources
  labeled_variants_conflicts.csv  — 仅 ≥2 标签的歧义变体
  foldx_a1_2B_clean.txt / foldx_a1_2M_clean.txt — A1 域内单标签变体(每行一个 aa_change)

用法:
  python3 scripts/pipeline/build_labeled_variant_set.py \
      --data-dir /Volumes/LQ1000/VWD --out-dir output
依赖: openpyxl (读 xlsx)。
"""
import argparse
import csv
import json
import re
from collections import defaultdict
from pathlib import Path

try:
    import openpyxl
except ImportError:
    raise SystemExit("[FATAL] 需要 openpyxl: pip install openpyxl")

AA3 = {'Ala': 'A', 'Arg': 'R', 'Asn': 'N', 'Asp': 'D', 'Cys': 'C', 'Gln': 'Q',
       'Glu': 'E', 'Gly': 'G', 'His': 'H', 'Ile': 'I', 'Leu': 'L', 'Lys': 'K',
       'Met': 'M', 'Phe': 'F', 'Pro': 'P', 'Ser': 'S', 'Thr': 'T', 'Trp': 'W',
       'Tyr': 'Y', 'Val': 'V'}
THREE = re.compile(r"\b([A-Z][a-z]{2})(\d{2,4})([A-Z][a-z]{2})\b")
JNAME = re.compile(r"VWF_([ACDEFGHIKLMNPQRSTVWY])(\d+)([ACDEFGHIKLMNPQRSTVWY])(?:_|$)")

# VWF domain ranges (UniProt P04275 成熟蛋白编号) — 用位置标注 domain
VWF_DOMAINS = [
    ("D'", 764, 865), ("D3", 866, 1241), ("A1", 1242, 1481), ("A2", 1482, 1672),
    ("A3", 1673, 1874), ("D4", 1875, 2255), ("C", 2256, 2577), ("CK", 2578, 2813),
    ("D1D2_pro", 1, 763),
]
A1_7A6O = (1262, 1466)   # 7A6O 实验结构解析范围 → FoldX 可直接 graft

FOLDER_LABEL = {"1型": "1", "2A型": "2A", "2B型": "2B", "2M型": "2M", "2N型": "2N", "3型": "3"}


def domain_of(pos: int) -> str:
    for name, lo, hi in VWF_DOMAINS:
        if lo <= pos <= hi:
            return name
    return "?"


def parse_xlsx(path):
    """只读主数据表 Sheet1 的 'Amino acid change' 列。

    坑: Sheet1 下半部分是参考文献(论文标题里嵌着 Arg273Trp 等突变名),各表共享同一
    文献集 → 若全表扫会把文献里的变体也算进来、造成假的跨型冲突。故:
      - 只读 'Amino acid change' 那一列;
      - 只接受**短**(<40 字符)且以 3 字母变体开头的格(排除文献长文本/说明行)。
    """
    out = set()
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Sheet1"] if "Sheet1" in wb.sheetnames else wb.active
    rows = list(ws.iter_rows(values_only=True))
    # 找 'Amino acid change' 列
    aac = None
    for r in rows[:8]:
        for ci, c in enumerate(r):
            if isinstance(c, str) and "amino acid" in c.lower():
                aac = ci
                break
        if aac is not None:
            break
    if aac is None:
        return out
    head = re.compile(r"^p?\.?\s*([A-Z][a-z]{2})(\d{2,4})([A-Z][a-z]{2})")
    for r in rows:
        if aac >= len(r):
            continue
        c = r[aac]
        if not isinstance(c, str):
            continue
        s = c.strip()
        if len(s) >= 40:        # 文献/长说明 → 排除
            continue
        m = head.match(s)       # 必须以变体开头 (排除 "Mostly described as..." 等)
        if m:
            w, p, mut = AA3.get(m.group(1)), int(m.group(2)), AA3.get(m.group(3))
            if w and mut and w != mut:
                out.add((w, p, mut))
    return out


def parse_json(path):
    out = set()
    try:
        d = json.load(open(path))
    except Exception:
        return out
    for it in (d if isinstance(d, list) else []):
        nm = it.get("name", "") if isinstance(it, dict) else ""
        m = JNAME.search(nm)
        if m and m.group(1) != m.group(3):
            out.add((m.group(1), int(m.group(2)), m.group(3)))
    return out


def main():
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--data-dir", default="/Volumes/LQ1000/VWD")
    ap.add_argument("--out-dir", default="output")
    args = ap.parse_args()

    root = Path(__file__).resolve().parent.parent.parent
    data = Path(args.data_dir)
    out = (root / args.out_dir) if not Path(args.out_dir).is_absolute() else Path(args.out_dir)
    out.mkdir(parents=True, exist_ok=True)
    if not data.is_dir():
        raise SystemExit(f"[FATAL] data-dir 不存在: {data}")

    # variant(单字母 aa_change) -> {label: set(source files)}
    vmap = defaultdict(lambda: defaultdict(set))
    for folder, label in FOLDER_LABEL.items():
        fdir = data / folder
        if not fdir.is_dir():
            continue
        files = [p for p in fdir.iterdir()
                 if p.is_file() and not p.name.startswith("._")]
        for f in files:
            if f.suffix.lower() == ".xlsx":
                variants = parse_xlsx(f)
            elif f.suffix.lower() == ".json":
                variants = parse_json(f)
            else:
                continue
            for (w, p, mut) in variants:
                vmap[(w, p, mut)][label].add(f.name)

    # 汇总成行
    rows = []
    for (w, p, mut), labmap in sorted(vmap.items(), key=lambda x: x[0][1]):
        labels = sorted(labmap.keys())
        rows.append({
            "aa_change": f"{w}{p}{mut}",
            "wt_aa": w, "position": p, "mut_aa": mut,
            "labels": "|".join(labels),
            "n_labels": len(labels),
            "conflict": len(labels) > 1,
            "domain": domain_of(p),
            "in_a1_7a6o": A1_7A6O[0] <= p <= A1_7A6O[1],
            "sources": ";".join(sorted({s for ss in labmap.values() for s in ss})),
        })

    cols = ["aa_change", "wt_aa", "position", "mut_aa", "labels", "n_labels",
            "conflict", "domain", "in_a1_7a6o", "sources"]
    with open(out / "labeled_variants_all.csv", "w", newline="") as fh:
        w_ = csv.DictWriter(fh, fieldnames=cols); w_.writeheader(); w_.writerows(rows)

    conflicts = [r for r in rows if r["conflict"]]
    with open(out / "labeled_variants_conflicts.csv", "w", newline="") as fh:
        w_ = csv.DictWriter(fh, fieldnames=cols); w_.writeheader(); w_.writerows(conflicts)

    # FoldX 直接可用清单: A1 域内 + 单标签(干净)
    def clean_a1(label):
        return sorted(r["aa_change"] for r in rows
                      if r["in_a1_7a6o"] and not r["conflict"] and r["labels"] == label)
    for label in ("2B", "2M", "2A"):
        vs = clean_a1(label)
        (out / f"foldx_a1_{label}_clean.txt").write_text("\n".join(vs) + ("\n" if vs else ""))

    # ---- 报告 ----
    from collections import Counter
    by_label = Counter()
    a1_clean = Counter()
    for r in rows:
        for lab in r["labels"].split("|"):
            by_label[lab] += 1
        if r["in_a1_7a6o"] and not r["conflict"]:
            a1_clean[r["labels"]] += 1
    print("=" * 60)
    print(f"总唯一变体: {len(rows)}  (冲突/多标签: {len(conflicts)})")
    print("各分型出现次数(含冲突重复计):", dict(by_label))
    print("\nA1 域(1262-1466)内、单标签(FoldX 干净可用):")
    for lab in ("2B", "2M", "2A", "1", "2N", "3"):
        if a1_clean.get(lab):
            print(f"  {lab}: {a1_clean[lab]}")
    print("\n冲突变体样例(同一变体多个分型标签):")
    for r in conflicts[:10]:
        print(f"  {r['aa_change']:>8}  labels={r['labels']:>8}  domain={r['domain']}")
    print("=" * 60)
    print(f"输出 → {out}/")
    print("  labeled_variants_all.csv / labeled_variants_conflicts.csv")
    print("  foldx_a1_{2B,2M,2A}_clean.txt")
    print(f"用法: build_2b_mutants_foldx.py --variants-file {out}/foldx_a1_2B_clean.txt")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
